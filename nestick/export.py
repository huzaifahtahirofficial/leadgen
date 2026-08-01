"""Export leads to CSV, JSON, JSONL, XLSX, Markdown and SQLite."""

from __future__ import annotations

import contextlib
import csv
import json
import sqlite3
import time
from pathlib import Path
from typing import Any, Iterable, Sequence

from .config import Settings
from .models import Lead, Stats
from .security import sanitise_cell
from .utils import log

COLUMNS: tuple[str, ...] = (
    "domain", "name", "score", "emails", "phones", "website", "address",
    "linkedin", "twitter", "facebook", "instagram", "youtube", "tiktok",
    "github", "telegram", "whatsapp", "medium",
    "rating", "reviews", "category", "latitude", "longitude",
    "title", "description", "source", "pages_crawled", "fetched_at",
)


def _row(lead: Lead) -> dict[str, Any]:
    socials = lead.socials
    return {
        "domain": lead.domain,
        "name": lead.name or "",
        "score": lead.score,
        "emails": "; ".join(lead.emails),
        "phones": "; ".join(lead.phones),
        "website": lead.url,
        "address": lead.address or "",
        **{k: "; ".join(socials.get(k, [])) for k in (
            "linkedin", "twitter", "facebook", "instagram", "youtube",
            "tiktok", "github", "telegram", "whatsapp", "medium")},
        "rating": lead.rating if lead.rating is not None else "",
        "reviews": lead.reviews if lead.reviews is not None else "",
        "category": lead.category or "",
        "latitude": lead.latitude if lead.latitude is not None else "",
        "longitude": lead.longitude if lead.longitude is not None else "",
        "title": lead.title or "",
        "description": lead.description or "",
        "source": lead.source,
        "pages_crawled": lead.pages_crawled,
        "fetched_at": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(lead.fetched_at)),
    }


class Exporter:
    """Writes a result set in every requested format, atomically."""

    def __init__(self, settings: Settings) -> None:
        self.s = settings
        self.base = Path(settings.output).expanduser()
        if self.base.suffix.lower() in {".csv", ".json", ".jsonl", ".xlsx", ".md", ".db"}:
            self.base = self.base.with_suffix("")
        self.base.parent.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------ #
    def write(self, leads: Sequence[Lead], stats: Stats | None = None) -> list[Path]:
        written: list[Path] = []
        formats = {f.lower() for f in self.s.formats}
        if "all" in formats:
            formats = {"csv", "json", "jsonl", "xlsx", "md", "sqlite"}
        dispatch = {
            "csv": self.csv, "json": self.json, "jsonl": self.jsonl,
            "xlsx": self.xlsx, "excel": self.xlsx, "md": self.markdown,
            "markdown": self.markdown, "sqlite": self.sqlite, "db": self.sqlite,
        }
        for fmt in formats:
            fn = dispatch.get(fmt)
            if not fn:
                log.warning("Unknown output format %r — skipped", fmt)
                continue
            try:
                p = fn(leads, stats)
                written.append(p)
                log.info("Wrote %s (%d leads)", p, len(leads))
            except Exception as exc:  # noqa: BLE001
                log.error("Export %s failed: %s", fmt, exc)
        return written

    # ------------------------------------------------------------------ #
    def csv(self, leads: Sequence[Lead], stats: Stats | None = None) -> Path:
        p = self.base.with_suffix(".csv")
        tmp = p.with_suffix(".csv.tmp")
        with tmp.open("w", newline="", encoding="utf-8-sig") as fh:
            w = csv.DictWriter(fh, fieldnames=list(COLUMNS), extrasaction="ignore")
            w.writeheader()
            for lead in leads:
                # Scraped text is attacker-controlled: neutralise spreadsheet
                # formulas ("=cmd|..." would execute when opened in Excel).
                w.writerow({k: sanitise_cell(v) for k, v in _row(lead).items()})
        tmp.replace(p)
        return p

    def json(self, leads: Sequence[Lead], stats: Stats | None = None) -> Path:
        p = self.base.with_suffix(".json")
        payload = {
            "generated_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
            "settings": self.s.to_dict(),
            "stats": stats.as_row() if stats else {},
            "count": len(leads),
            "leads": [l.to_dict() for l in leads],
        }
        try:
            import orjson

            data = orjson.dumps(payload, option=orjson.OPT_INDENT_2)
            p.write_bytes(data)
        except Exception:
            p.write_text(json.dumps(payload, indent=2, default=str), "utf-8")
        return p

    def jsonl(self, leads: Sequence[Lead], stats: Stats | None = None) -> Path:
        p = self.base.with_suffix(".jsonl")
        with p.open("w", encoding="utf-8") as fh:
            for lead in leads:
                fh.write(json.dumps(lead.to_dict(), default=str, ensure_ascii=False) + "\n")
        return p

    def xlsx(self, leads: Sequence[Lead], stats: Stats | None = None) -> Path:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        p = self.base.with_suffix(".xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        ws.append(list(COLUMNS))

        header_fill = PatternFill("solid", fgColor="1F3B57")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

        for lead in leads:
            row = _row(lead)
            ws.append([sanitise_cell(row.get(c, "")) for c in COLUMNS])

        # Hyperlink websites + colour-code the score column.
        score_col = COLUMNS.index("score") + 1
        site_col = COLUMNS.index("website") + 1
        for r in range(2, ws.max_row + 1):
            c = ws.cell(row=r, column=site_col)
            if isinstance(c.value, str) and c.value.startswith("http"):
                c.hyperlink = c.value
                c.font = Font(color="0563C1", underline="single")
            sc = ws.cell(row=r, column=score_col)
            if isinstance(sc.value, (int, float)):
                colour = "C6EFCE" if sc.value >= 60 else "FFEB9C" if sc.value >= 30 else "FFC7CE"
                sc.fill = PatternFill("solid", fgColor=colour)

        widths = {"domain": 28, "name": 30, "emails": 46, "phones": 24,
                  "website": 38, "address": 40, "description": 60, "title": 40}
        for i, col in enumerate(COLUMNS, 1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 16)

        if ws.max_row > 1:
            ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
            table = Table(displayName="Leads", ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True
            )
            with contextlib.suppress(Exception):
                ws.add_table(table)

        # Summary sheet
        s2 = wb.create_sheet("Summary")
        s2.append(["Metric", "Value"])
        s2["A1"].font = s2["B1"].font = Font(bold=True)
        rows: list[tuple[str, Any]] = [
            ("Leads", len(leads)),
            ("With e-mail", sum(1 for l in leads if l.emails)),
            ("With phone", sum(1 for l in leads if l.phones)),
            ("Total e-mails", sum(len(l.emails) for l in leads)),
            ("Avg score", round(sum(l.score for l in leads) / max(len(leads), 1), 1)),
        ]
        if stats:
            rows += [(k.replace("_", " ").title(), v) for k, v in stats.as_row().items()]
        for k, v in rows:
            s2.append([k, v])
        s2.column_dimensions["A"].width = 24
        s2.column_dimensions["B"].width = 18

        wb.save(p)
        return p

    def markdown(self, leads: Sequence[Lead], stats: Stats | None = None) -> Path:
        p = self.base.with_suffix(".md")
        lines = [
            f"# Lead report — {time.strftime('%Y-%m-%d %H:%M')}",
            "",
            f"**{len(leads)} leads** · "
            f"{sum(1 for l in leads if l.emails)} with e-mail · "
            f"{sum(len(l.emails) for l in leads)} addresses total",
            "",
            "| # | Domain | Name | Score | E-mails | Phones |",
            "|--:|---|---|--:|---|---|",
        ]
        for i, l in enumerate(leads, 1):
            lines.append(
                f"| {i} | [{l.domain}]({l.url or '#'}) | {(l.name or '')[:40]} | "
                f"{l.score} | {'<br>'.join(l.emails[:4]) or '—'} | "
                f"{'<br>'.join(l.phones[:2]) or '—'} |"
            )
        if stats:
            lines += ["", "## Run statistics", "", "| Metric | Value |", "|---|--:|"]
            lines += [f"| {k.replace('_', ' ')} | {v} |" for k, v in stats.as_row().items()]
        p.write_text("\n".join(lines) + "\n", "utf-8")
        return p

    def sqlite(self, leads: Sequence[Lead], stats: Stats | None = None) -> Path:
        p = self.base.with_suffix(".db")
        db = sqlite3.connect(p)
        try:
            db.executescript(
                """
                CREATE TABLE IF NOT EXISTS leads (
                    domain TEXT PRIMARY KEY, name TEXT, url TEXT, score REAL,
                    address TEXT, latitude REAL, longitude REAL, rating REAL,
                    reviews INTEGER, category TEXT, title TEXT, description TEXT,
                    source TEXT, pages_crawled INTEGER, fetched_at REAL
                );
                CREATE TABLE IF NOT EXISTS contacts (
                    domain TEXT, kind TEXT, value TEXT, confidence REAL,
                    source_url TEXT, meta TEXT,
                    PRIMARY KEY (domain, kind, value)
                );
                CREATE INDEX IF NOT EXISTS contacts_kind ON contacts(kind);
                """
            )
            db.executemany(
                "INSERT OR REPLACE INTO leads VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                [
                    (l.domain, l.name, l.url, l.score, l.address, l.latitude,
                     l.longitude, l.rating, l.reviews, l.category, l.title,
                     l.description, l.source, l.pages_crawled, l.fetched_at)
                    for l in leads
                ],
            )
            db.executemany(
                "INSERT OR REPLACE INTO contacts VALUES (?,?,?,?,?,?)",
                [
                    (l.domain, str(c.kind), c.value, c.confidence, c.source_url,
                     json.dumps(c.meta, default=str))
                    for l in leads for c in l.contacts
                ],
            )
            db.commit()
        finally:
            db.close()
        return p


def summarise(leads: Iterable[Lead]) -> dict[str, Any]:
    """Aggregate metrics used by the CLI summary panel."""
    leads = list(leads)
    emails = [e for l in leads for e in l.emails]
    return {
        "leads": len(leads),
        "with_email": sum(1 for l in leads if l.emails),
        "with_phone": sum(1 for l in leads if l.phones),
        "with_social": sum(1 for l in leads if l.socials),
        "emails": len(emails),
        "unique_emails": len(set(emails)),
        "avg_score": round(sum(l.score for l in leads) / max(len(leads), 1), 1),
    }
